import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from pathlib import Path
from shutil import copyfile

# Define paths for your input and output Excel files
original_file = 'C:\\Fackelmann_Internship\\SAP\\data.xlsx'
new_file = 'C:\\Fackelmann_Internship\\SAP\\new_datag.xlsx'

# Copy the original file
copyfile(original_file, new_file)

# Read the data from the CSV file
df = pd.read_csv(r"C:\Fackelmann_Internship\SAP\data.csv", header=None, encoding='utf-8', thousands=',')

# Extract the customer names
customers = df.iloc[12::, 1]
customers = customers[customers.notna()]
customers = customers.values

# Extract the metrics for each customer for each month
metrics = []
for j in range(len(customers)):
    customer_metrics = []
    for i in range(11):  # Change here to 11 months
        start_row = 12 + j * 1
        end_row = start_row + 1
        start_col = 2 + i * 4
        end_col = start_col + 4
        month_metrics = df.iloc[start_row:end_row, start_col:end_col].values.flatten()  # Flatten the 2D array to 1D
        customer_metrics.append(month_metrics)

    metrics.append(customer_metrics)

# Extract the starting fiscal year from the CSV file
start_fiscal_year = int(df.iloc[7, 2][-4:])

# Create a new DataFrame with the correct customer names and metrics
df = pd.DataFrame(np.concatenate(metrics, axis=0),
                  columns=['Gross Turnover', 'COGS', 'Total Sales Qty', 'Factor Gross Turnover / COGS'])

df['Customer'] = np.repeat(customers, 11)  # Change here to 11 months
df['Month'] = np.tile(np.arange(1, 12), len(customers))  # Change here to 11 months
# Add 'Fiscal Year' column
# Calculate fiscal year based on month and starting fiscal year
df['Fiscal Year'] = df['Month'].apply(lambda month: start_fiscal_year if 6 > month <= 12 else start_fiscal_year + 1)

# Convert the 'Value' column to numeric
for col in ['Gross Turnover', 'COGS', 'Total Sales Qty', 'Factor Gross Turnover / COGS']:
    df[col] = df[col].str.replace(',', '').astype(float)

# Reshape the data into a long format
df_long = df.melt(id_vars=['Customer', 'Month'], var_name='Metric', value_name='Value')

# Define month names
month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']  # Remove 'Jul'

# Load the copied Excel file you want to save the images in
book = load_workbook(new_file)
writer = pd.ExcelWriter(new_file, engine='openpyxl')
writer.book = book

for metric in ['Gross Turnover', 'COGS', 'Total Sales Qty', 'Factor Gross Turnover / COGS']:
    unique_customers = df_long['Customer'].unique()
    #print(unique_customers)
    num_batches = len(unique_customers) // 5 + (len(unique_customers) % 5 > 0)

    for batch in range(num_batches):
        plt.figure(figsize=(10, 6))

        start = batch * 5
        end = start + 5
        batch_customers = unique_customers[start:end]

        for customer in batch_customers:
            customer_df = df_long[df_long['Customer'] == customer]
            metric_df = customer_df[customer_df['Metric'] == metric]
            plt.plot(metric_df['Month'], metric_df['Value'].values, label=customer)

        plt.title(f'Progress of {metric} over time (Customers {start + 1}-{min(end, len(unique_customers))})')
        plt.xlabel('Month')
        plt.ylabel('Value')
        plt.xticks(np.arange(1, 12), month_names)  # Set the x-ticks to be the month names
        plt.legend()
        plt.grid(which='both', axis='both')  # Add this line to show a more detailed grid
        # Instead of plt.show(), save the figure into a BytesIO object
        image_stream = BytesIO()
        plt.savefig(image_stream, format='png')
        plt.close()

        # Seek back to the start of the BytesIO object
        image_stream.seek(0)

        # Create an Image object from the BytesIO stream and add it to the Excel file
        image = Image(image_stream)
        # Replace invalid characters in sheet title
        safe_title = f'{metric}_{batch + 1}'.replace('/', '_')
        sheet = writer.book.create_sheet(title=safe_title)
        sheet.add_image(image, 'A1')

# Save the changes to the Excel file
writer.save()

# Save the DataFrame df into a new Excel file
df.to_excel('C:\\Fackelmann_Internship\\SAP\\new_data.xlsx', index=False)